import openpyxl
import json
import os

wb = openpyxl.load_workbook(r'D:\feishu download\2026年整车试制项目运营表-7.06.xlsx', data_only=True)

# === 1. Extract assembly areas (场地) ===
ws_areas = wb['各装配场地资源安排一览图']
areas = []
for row in ws_areas.iter_rows(min_row=3, max_row=10, values_only=True):
    if row[0] and str(row[0]).strip() not in ('符号', 'None', ''):
        areas.append({
            'id': str(row[0]).strip(),
            'name': str(row[1]).strip() if row[1] else '',
            'lifts': str(row[2]).strip() if row[2] else '',
            'capacity': str(row[3]).strip() if row[3] else '',
            'monthly_data': {}
        })

# === 2. Extract weekly resource map (未来2个月) ===
ws_map = wb['未来2个月资源map图']
projects = []
current_project = None

for row in ws_map.iter_rows(min_row=5, max_row=ws_map.max_row, values_only=True):
    if row[0] and str(row[0]).strip():
        current_area = str(row[0]).strip()
    
    if row[4] and str(row[4]).strip() and str(row[4]) != '项目':
        proj_name = str(row[4]).strip()
        proj_type = str(row[5]).strip() if row[5] else ''
        proj_category = str(row[6]).strip() if row[6] else ''
        
        if proj_type in ('A', 'B', 'C') or proj_type in ('骡子车', 'ET0', '软模车'):
            category_map = {'A': 'A类', 'B': 'B类', 'C': 'C类', '骡子车': 'C类', 'ET0': 'B类', '软模车': 'A类'}
            category = category_map.get(proj_type, proj_type)
            
            total_qty = row[8] if row[8] else 0
            if isinstance(total_qty, (int, float)):
                total_qty = int(total_qty)
            else:
                total_qty = 0
            
            projects.append({
                'name': proj_name,
                'code': proj_name,
                'category': category,
                'type': '试制' if category in ('A类', 'B类') else '试制',
                'totalQty': total_qty,
                'summerQty': row[9] if row[9] else 0,
                'status': '进行中' if total_qty > 0 else '待定',
                'location': current_area if current_area else '',
                'weeks': {
                    'w1': str(row[12]) if row[12] else '',
                    'w2': str(row[13]) if row[13] else '',
                    'w3': str(row[14]) if row[14] else '',
                    'w4': str(row[15]) if row[15] else '',
                    'w5': str(row[16]) if row[16] else '',
                    'w6': str(row[17]) if row[17] else '',
                    'w7': str(row[18]) if row[18] else '',
                    'w8': str(row[19]) if row[19] else '',
                    'w9': str(row[20]) if row[20] else '',
                    'w10': str(row[21]) if row[21] else '',
                }
            })

# === 3. Extract project list (2026年) ===
ws_proj = wb['2026年各项目需求及装配资源-试制主管&各组长']
project_list = []
for row in ws_proj.iter_rows(min_row=4, max_row=min(170, ws_proj.max_row), values_only=True):
    if row[1] and str(row[1]).strip() not in ('None', ''):
        proj_name = str(row[1]).strip()
        proj_cat = str(row[5]).strip() if row[5] else ''
        pm = str(row[11]).strip() if row[11] else ''
        
        cat_map = {'A类': 'A类', 'B类': 'B类', 'C类': 'C类', '等同C类装配': 'C类'}
        category = cat_map.get(proj_cat, 'C类')
        
        # Monthly data (columns 47-58 = Jan-Dec 2026)
        monthly = {}
        months = ['1月', '2月', '3月', '4月', '5月', '6月', '7月', '8月', '9月', '10月', '11月', '12月']
        for i, m in enumerate(months):
            col_idx = 47 + i
            if col_idx < len(row):
                val = row[col_idx]
                if val and str(val).strip() not in ('None', ''):
                    monthly[m] = str(val).strip()
        
        if proj_name and proj_name != 'None':
            project_list.append({
                'name': proj_name,
                'category': category,
                'pm': pm,
                'monthly': monthly
            })

# === Output summary ===
print(f'装配场地: {len(areas)}个')
for a in areas:
    print(f'  {a["id"]}: {a["name"]} | {a["lifts"]}')

print(f'\n未来2个月项目: {len(projects)}个')
print(f'2026年项目列表: {len(project_list)}个')

# Generate output data
output = {
    'areas': areas,
    'projects': projects,
    'projectList': project_list
}

# Save as JSON for reference
out_path = os.path.join(os.path.dirname(__file__), '..', 'assets', 'data', 'extracted_data.json')
os.makedirs(os.path.dirname(out_path), exist_ok=True)
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2, default=str)

print(f'\nData saved to: {out_path}')
